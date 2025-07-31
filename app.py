from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, send_file
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
import json
import os
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import threading

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production'

# Flask-Mail configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'your-email@gmail.com'  # Change this
app.config['MAIL_PASSWORD'] = 'your-app-password'     # Change this
app.config['MAIL_DEFAULT_SENDER'] = 'your-email@gmail.com'

mail = Mail(app)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'

# Simple User class
class User(UserMixin):
    def __init__(self, id, username, password_hash):
        self.id = id
        self.username = username
        self.password_hash = password_hash

@login_manager.user_loader
def load_user(user_id):
    data = load_data()
    user_data = next((u for u in data.get('users', []) if u['id'] == int(user_id)), None)
    if user_data:
        return User(user_data['id'], user_data['username'], user_data['password_hash'])
    return None

# Due date helper functions
def get_due_date_class(due_date_str):
    if not due_date_str:
        return ''
    
    try:
        due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        today = date.today()
        
        if due_date < today:
            return 'card-overdue'
        elif due_date == today:
            return 'card-due-today'
        elif (due_date - today).days <= 3:
            return 'card-due-soon'
        return ''
    except:
        return ''

def get_due_date_text_class(due_date_str):
    if not due_date_str:
        return ''
    
    try:
        due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        today = date.today()
        
        if due_date < today:
            return 'due-date-overdue'
        elif due_date == today:
            return 'due-date-today'
        return 'due-date-upcoming'
    except:
        return ''

def get_due_date_status(due_date_str):
    if not due_date_str:
        return ''
    
    try:
        due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        today = date.today()
        
        if due_date < today:
            return '(OVERDUE)'
        elif due_date == today:
            return '(TODAY)'
        elif (due_date - today).days <= 3:
            return '(SOON)'
        return ''
    except:
        return ''

def get_comment_count(card_id):
    data = load_data()
    return len([c for c in data.get('comments', []) if c['card_id'] == card_id])

def format_timestamp(timestamp):
    """Format timestamp to YYYY-MM-DD HH:MM:SS (remove milliseconds)"""
    if not timestamp:
        return '-'
    
    # If it's already in the desired format, return as is
    if isinstance(timestamp, str) and len(timestamp) == 10:  # YYYY-MM-DD format
        return timestamp
    
    # Try to parse and format
    try:
        if isinstance(timestamp, str):
            # Try parsing different formats
            for fmt in ['%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                try:
                    dt = datetime.strptime(timestamp, fmt)
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                except ValueError:
                    continue
        return timestamp  # Return original if parsing fails
    except:
        return timestamp  # Return original if any error occurs

def send_overdue_notifications():
    """Send email notifications for overdue cards"""
    try:
        data = load_data()
        today = date.today()
        
        # Find overdue cards
        overdue_cards = []
        for card in data['cards']:
            if card.get('due_date'):
                try:
                    due_date = datetime.strptime(card['due_date'], '%Y-%m-%d').date()
                    if due_date < today and card['status'] != 'done':
                        overdue_cards.append(card)
                except:
                    continue
        
        if not overdue_cards:
            return
        
        # Get users with email notifications enabled
        users_to_notify = []
        for user in data.get('users', []):
            if user.get('email_notifications', True):  # Default to True
                users_to_notify.append(user)
        
        # Send emails
        for user in users_to_notify:
            if user.get('email'):  # Only if user has email set
                send_overdue_email(user, overdue_cards)
                
    except Exception as e:
        print(f"Error sending notifications: {e}")

def send_overdue_email(user, overdue_cards):
    """Send overdue cards email to a specific user"""
    try:
        projects = {p['id']: p['name'] for p in load_data()['projects']}
        
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background: #0079bf; color: white; padding: 20px; text-align: center;">
                <h1>🚨 Overdue Cards Alert</h1>
            </div>
            
            <div style="padding: 20px;">
                <p>Hello {user['username']},</p>
                <p>You have <strong>{len(overdue_cards)} overdue cards</strong> that need attention:</p>
                
                <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 20px 0;">
        """
        
        for card in overdue_cards:
            project_name = projects.get(card['project_id'], 'Unknown')
            html_body += f"""
                    <div style="border-left: 4px solid #ff5630; padding: 10px; margin: 10px 0; background: white;">
                        <h3 style="margin: 0 0 5px 0; color: #172b4d;">#{card['id']} - {card['title']}</h3>
                        <p style="margin: 5px 0; color: #5e6c84;">Project: {project_name}</p>
                        <p style="margin: 5px 0; color: #5e6c84;">Due: {card['due_date']} | Priority: {card['priority']} | Assignee: {card.get('assignee', 'Unassigned')}</p>
                        <p style="margin: 5px 0;">{card.get('description', '')}</p>
                    </div>
            """
        
        html_body += f"""
                </div>
                
                <p>Please review and update these cards as soon as possible.</p>
                
                <div style="text-align: center; margin: 30px 0;">
                    <a href="http://localhost:5000/issues" 
                       style="background: #0079bf; color: white; padding: 12px 24px; text-decoration: none; border-radius: 6px; display: inline-block;">
                        View All Issues
                    </a>
                </div>
                
                <hr style="margin: 30px 0; border: none; border-top: 1px solid #eee;">
                <p style="font-size: 12px; color: #5e6c84; text-align: center;">
                    This is an automated notification from PM Tool.<br>
                    Generated on {datetime.now().strftime('%Y-%m-%d at %H:%M')}
                </p>
            </div>
        </body>
        </html>
        """
        
        msg = Message(
            subject=f"PM Tool: {len(overdue_cards)} Overdue Cards Need Attention",
            recipients=[user['email']],
            html=html_body
        )
        
        mail.send(msg)
        print(f"Overdue notification sent to {user['username']}")
        
    except Exception as e:
        print(f"Error sending email to {user['username']}: {e}")

def schedule_daily_notifications():
    """Schedule daily email notifications (runs in background)"""
    def run_notifications():
        import time
        while True:
            # Check time - send at 9 AM daily
            now = datetime.now()
            if now.hour == 9 and now.minute == 0:
                send_overdue_notifications()
                time.sleep(60)  # Wait 1 minute to avoid duplicate sends
            time.sleep(30)  # Check every 30 seconds
    
    # Run in background thread
    thread = threading.Thread(target=run_notifications, daemon=True)
    thread.start()

# Register template filters
app.jinja_env.globals.update(
    get_due_date_class=get_due_date_class,
    get_due_date_text_class=get_due_date_text_class,
    get_due_date_status=get_due_date_status,
    get_comment_count=get_comment_count,
    format_timestamp=format_timestamp
)

DATA_FILE = 'data.json'

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
            # Ensure users section exists
            if 'users' not in data:
                data['users'] = [
                    {'id': 1, 'username': 'admin', 'password_hash': generate_password_hash('admin123')}
                ]
                save_data(data)
            # Ensure comments section exists
            if 'comments' not in data:
                data['comments'] = []
                save_data(data)
            return data
    return {
        'users': [
            {'id': 1, 'username': 'admin', 'password_hash': generate_password_hash('admin123')}
        ],
        'projects': [{
            'id': 1,
            'name': 'Sample Project',
            'description': 'Your first project'
        }],
        'cards': [
            {'id': 1, 'project_id': 1, 'title': 'Setup project', 'description': 'Initial setup', 'status': 'done', 'assignee': 'John', 'priority': 'High', 'created_at': '2024-01-01', 'due_date': '2024-01-15'},
            {'id': 2, 'project_id': 1, 'title': 'Design UI', 'description': 'Create mockups', 'status': 'in_progress', 'assignee': 'Jane', 'priority': 'Medium', 'created_at': '2024-01-02', 'due_date': '2024-12-30'},
            {'id': 3, 'project_id': 1, 'title': 'Implement backend', 'description': 'API development', 'status': 'todo', 'assignee': 'Bob', 'priority': 'High', 'created_at': '2024-01-03', 'due_date': '2024-12-20'}
        ],
        'comments': [
            {'id': 1, 'card_id': 1, 'author': 'admin', 'content': 'Great work on the setup!', 'created_at': '2024-01-02 10:30:00'},
            {'id': 2, 'card_id': 2, 'author': 'admin', 'content': 'Please update the color scheme', 'created_at': '2024-01-03 14:15:00'}
        ]
    }

def save_data(data):
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=2)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        data = load_data()
        user_data = next((u for u in data.get('users', []) if u['username'] == username), None)
        
        if user_data and check_password_hash(user_data['password_hash'], password):
            user = User(user_data['id'], user_data['username'], user_data['password_hash'])
            login_user(user)
            return redirect(url_for('home'))
        else:
            flash('Invalid username or password')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        data = load_data()
        
        # Check if user exists
        if any(u['username'] == username for u in data.get('users', [])):
            flash('Username already exists')
            return render_template('register.html')
        
        # Create new user
        new_user = {
            'id': max([u['id'] for u in data.get('users', [])], default=0) + 1,
            'username': username,
            'password_hash': generate_password_hash(password)
        }
        
        if 'users' not in data:
            data['users'] = []
        data['users'].append(new_user)
        save_data(data)
        
        flash('Registration successful! Please log in.')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    data = load_data()
    user_data = next((u for u in data['users'] if u['id'] == current_user.id), None)
    
    if request.method == 'POST':
        # Update user email and notification preferences
        email = request.form.get('email', '')
        email_notifications = 'email_notifications' in request.form
        
        for user in data['users']:
            if user['id'] == current_user.id:
                user['email'] = email
                user['email_notifications'] = email_notifications
                break
        
        save_data(data)
        flash('Profile updated successfully!')
        return redirect(url_for('profile'))
    
    return render_template('profile.html', user=user_data)

@app.route('/api/test-email')
@login_required
def test_email():
    """Test email functionality"""
    try:
        data = load_data()
        user_data = next((u for u in data['users'] if u['id'] == current_user.id), None)
        
        if not user_data or not user_data.get('email'):
            return jsonify({'error': 'No email address set in profile'}), 400
        
        # Send test email
        msg = Message(
            subject="PM Tool - Test Email",
            recipients=[user_data['email']],
            html="""
            <h2>Test Email Successful! 🎉</h2>
            <p>Your email notifications are working correctly.</p>
            <p>You will receive daily notifications for overdue cards.</p>
            """
        )
        
        mail.send(msg)
        return jsonify({'success': True, 'message': 'Test email sent successfully!'})
        
    except Exception as e:
        return jsonify({'error': f'Email test failed: {str(e)}'}), 500

@app.route('/')
@login_required
def home():
    return render_template('home.html')

@app.route('/dashboard')
@login_required
def dashboard():
    data = load_data()
    all_cards = data['cards']
    stats = {
        'todo': len([c for c in all_cards if c['status'] == 'todo']),
        'in_progress': len([c for c in all_cards if c['status'] == 'in_progress']),
        'done': len([c for c in all_cards if c['status'] == 'done']),
        'total': len(all_cards)
    }
    return render_template('dashboard.html', projects=data['projects'], stats=stats)

@app.route('/issues')
@login_required
def issues_list():
    data = load_data()
    project_id = request.args.get('project_id', type=int)
    
    if project_id:
        cards = [c for c in data['cards'] if c['project_id'] == project_id]
        project = next((p for p in data['projects'] if p['id'] == project_id), None)
    else:
        cards = data['cards']
        project = None
    
    return render_template('issues.html', cards=cards, projects=data['projects'], current_project=project)

@app.route('/board/<int:project_id>')
@login_required
def kanban_board(project_id):
    data = load_data()
    project = next((p for p in data['projects'] if p['id'] == project_id), None)
    if not project:
        return "Project not found", 404
    
    cards = [c for c in data['cards'] if c['project_id'] == project_id]
    todo_cards = [c for c in cards if c['status'] == 'todo']
    in_progress_cards = [c for c in cards if c['status'] == 'in_progress']
    done_cards = [c for c in cards if c['status'] == 'done']
    
    return render_template('kanban.html', 
                         project=project,
                         todo_cards=todo_cards,
                         in_progress_cards=in_progress_cards,
                         done_cards=done_cards)

@app.route('/api/move_card', methods=['POST'])
def move_card():
    data = load_data()
    card_id = request.json['card_id']
    new_status = request.json['status']
    
    for card in data['cards']:
        if card['id'] == card_id:
            card['status'] = new_status
            break
    
    save_data(data)
    return jsonify({'success': True})

@app.route('/api/add_card', methods=['POST'])
def add_card():
    data = load_data()
    new_card = {
        'id': max([c['id'] for c in data['cards']], default=0) + 1,
        'project_id': request.json['project_id'],
        'title': request.json['title'],
        'description': request.json.get('description', ''),
        'status': 'todo',
        'assignee': request.json.get('assignee', ''),
        'priority': request.json.get('priority', 'Medium'),
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'due_date': request.json.get('due_date', ''),
        'story_points': request.json.get('story_points')
    }
    data['cards'].append(new_card)
    save_data(data)
    return jsonify(new_card)

@app.route('/api/add_project', methods=['POST'])
def add_project():
    data = load_data()
    new_project = {
        'id': max([p['id'] for p in data['projects']], default=0) + 1,
        'name': request.json['name'],
        'description': request.json.get('description', '')
    }
    data['projects'].append(new_project)
    save_data(data)
    return jsonify(new_project)

@app.route('/api/update_card_status', methods=['POST'])
def update_card_status():
    data = load_data()
    card_id = request.json['card_id']
    new_status = request.json['status']
    
    for card in data['cards']:
        if card['id'] == card_id:
            card['status'] = new_status
            break
    
    save_data(data)
    return jsonify({'success': True})

@app.route('/api/card/<int:card_id>/comments')
@login_required
def get_comments(card_id):
    data = load_data()
    comments = [c for c in data.get('comments', []) if c['card_id'] == card_id]
    return jsonify(comments)

@app.route('/api/card/<int:card_id>/comments', methods=['POST'])
@login_required
def add_comment(card_id):
    data = load_data()
    new_comment = {
        'id': max([c['id'] for c in data.get('comments', [])], default=0) + 1,
        'card_id': card_id,
        'author': current_user.username,
        'content': request.json['content'],
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    if 'comments' not in data:
        data['comments'] = []
    data['comments'].append(new_comment)
    save_data(data)
    return jsonify(new_comment)

@app.route('/api/card/<int:card_id>')
@login_required
def get_card_details(card_id):
    data = load_data()
    card = next((c for c in data['cards'] if c['id'] == card_id), None)
    if not card:
        return jsonify({'error': 'Card not found'}), 404
    
    comments = [c for c in data.get('comments', []) if c['card_id'] == card_id]
    return jsonify({'card': card, 'comments': comments})

@app.route('/api/search')
@login_required
def search_cards():
    data = load_data()
    cards = data['cards']
    
    # Get search parameters
    search_query = request.args.get('q', '').lower()
    status_filter = request.args.get('status', '')
    priority_filter = request.args.get('priority', '')
    assignee_filter = request.args.get('assignee', '')
    project_filter = request.args.get('project_id', '')
    due_date_filter = request.args.get('due_date_filter', '')
    sort_by = request.args.get('sort', 'created_desc')
    
    # Filter cards
    filtered_cards = cards
    
    # Project filter
    if project_filter:
        filtered_cards = [c for c in filtered_cards if c['project_id'] == int(project_filter)]
    
    # Status filter
    if status_filter:
        filtered_cards = [c for c in filtered_cards if c['status'] == status_filter]
    
    # Priority filter
    if priority_filter:
        filtered_cards = [c for c in filtered_cards if c['priority'] == priority_filter]
    
    # Assignee filter
    if assignee_filter:
        if assignee_filter == '':
            filtered_cards = [c for c in filtered_cards if not c.get('assignee')]
        else:
            filtered_cards = [c for c in filtered_cards if c.get('assignee', '').lower() == assignee_filter.lower()]
    
    # Due date filter
    if due_date_filter:
        from datetime import datetime, date, timedelta
        today = date.today()
        
        if due_date_filter == 'overdue':
            filtered_cards = [c for c in filtered_cards if c.get('due_date') and 
                            datetime.strptime(c['due_date'], '%Y-%m-%d').date() < today]
        elif due_date_filter == 'today':
            filtered_cards = [c for c in filtered_cards if c.get('due_date') and 
                            datetime.strptime(c['due_date'], '%Y-%m-%d').date() == today]
        elif due_date_filter == 'this_week':
            week_end = today + timedelta(days=7)
            filtered_cards = [c for c in filtered_cards if c.get('due_date') and 
                            today <= datetime.strptime(c['due_date'], '%Y-%m-%d').date() <= week_end]
        elif due_date_filter == 'next_week':
            next_week_start = today + timedelta(days=7)
            next_week_end = today + timedelta(days=14)
            filtered_cards = [c for c in filtered_cards if c.get('due_date') and 
                            next_week_start <= datetime.strptime(c['due_date'], '%Y-%m-%d').date() <= next_week_end]
        elif due_date_filter == 'no_date':
            filtered_cards = [c for c in filtered_cards if not c.get('due_date')]
    
    # Search query (title and description)
    if search_query:
        filtered_cards = [c for c in filtered_cards if 
                         search_query in c['title'].lower() or 
                         search_query in c.get('description', '').lower()]
    
    # Sort cards
    if sort_by == 'created_desc':
        filtered_cards.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    elif sort_by == 'created_asc':
        filtered_cards.sort(key=lambda x: x.get('created_at', ''))
    elif sort_by == 'priority':
        priority_order = {'High': 3, 'Medium': 2, 'Low': 1}
        filtered_cards.sort(key=lambda x: priority_order.get(x.get('priority', 'Low'), 1), reverse=True)
    elif sort_by == 'due_date':
        # Sort by due date, putting items without due dates at the end
        def due_date_sort_key(card):
            if not card.get('due_date'):
                return '9999-12-31'  # Far future date for items without due dates
            return card['due_date']
        filtered_cards.sort(key=due_date_sort_key)
    elif sort_by == 'title':
        filtered_cards.sort(key=lambda x: x.get('title', '').lower())
    
    return jsonify({
        'cards': filtered_cards,
        'total': len(filtered_cards)
    })

@app.route('/api/move_to_backlog', methods=['POST'])
@login_required
def move_to_backlog():
    try:
        data = load_data()
        card_id = request.json['card_id']
        
        # Find and update the card status to 'todo' (backlog)
        for card in data['cards']:
            if card['id'] == card_id:
                card['status'] = 'todo'
                break
        else:
            return jsonify({'success': False, 'error': 'Card not found'}), 404
        
        save_data(data)
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/send_to_assignee', methods=['POST'])
@login_required
def send_to_assignee():
    try:
        data = load_data()
        card_id = request.json['card_id']
        assignee = request.json['assignee']
        
        # Find the card
        card = next((c for c in data['cards'] if c['id'] == card_id), None)
        if not card:
            return jsonify({'success': False, 'error': 'Card not found'}), 404
        
        # Find project name
        project = next((p for p in data['projects'] if p['id'] == card['project_id']), None)
        project_name = project['name'] if project else 'Unknown Project'
        
        # Send email notification (simplified version)
        try:
            msg = Message(
                subject=f"PM Tool: Issue #{card['id']} - {card['title']}",
                recipients=[f"{assignee}@company.com"],  # You may want to store actual emails
                html=f"""
                <h2>Issue Assignment Notification</h2>
                <p>You have been assigned to work on the following issue:</p>
                <div style="border: 1px solid #ddd; padding: 15px; margin: 10px 0;">
                    <h3>#{card['id']} - {card['title']}</h3>
                    <p><strong>Project:</strong> {project_name}</p>
                    <p><strong>Priority:</strong> {card['priority']}</p>
                    <p><strong>Status:</strong> {card['status'].replace('_', ' ').title()}</p>
                    <p><strong>Description:</strong> {card.get('description', 'No description')}</p>
                    {f"<p><strong>Due Date:</strong> {card['due_date']}</p>" if card.get('due_date') else ""}
                </div>
                <p>Please review and update the issue status as needed.</p>
                <a href="http://localhost:5000/issues" style="background: #0079bf; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px;">View Issues</a>
                """
            )
            mail.send(msg)
            return jsonify({'success': True})
        except Exception as email_error:
            return jsonify({'success': False, 'error': f'Failed to send email: {str(email_error)}'}), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/delete_card', methods=['POST'])
@login_required
def delete_card():
    try:
        data = load_data()
        card_id = request.json['card_id']
        
        # Find and remove the card
        original_count = len(data['cards'])
        data['cards'] = [c for c in data['cards'] if c['id'] != card_id]
        
        if len(data['cards']) == original_count:
            return jsonify({'success': False, 'error': 'Card not found'}), 404
        
        # Also remove associated comments
        data['comments'] = [c for c in data.get('comments', []) if c['card_id'] != card_id]
        
        save_data(data)
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/update_story_points', methods=['POST'])
@login_required
def update_story_points():
    try:
        data = load_data()
        card_id = request.json['card_id']
        story_points = request.json.get('story_points')
        
        # Find and update the card
        for card in data['cards']:
            if card['id'] == card_id:
                card['story_points'] = story_points
                break
        else:
            return jsonify({'success': False, 'error': 'Card not found'}), 404
        
        save_data(data)
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/move_to_sprint', methods=['POST'])
@login_required
def move_to_sprint():
    try:
        data = load_data()
        card_id = request.json['card_id']
        
        # Find and update the card status to 'in_progress' (active sprint)
        for card in data['cards']:
            if card['id'] == card_id:
                card['status'] = 'in_progress'
                break
        else:
            return jsonify({'success': False, 'error': 'Card not found'}), 404
        
        save_data(data)
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export-excel')
@login_required
def export_excel():
    data = load_data()
    cards = data['cards']
    projects = {p['id']: p['name'] for p in data['projects']}
    
    # Apply same filters as search
    search_query = request.args.get('q', '').lower()
    status_filter = request.args.get('status', '')
    priority_filter = request.args.get('priority', '')
    assignee_filter = request.args.get('assignee', '')
    project_filter = request.args.get('project_id', '')
    
    # Filter cards (same logic as search)
    filtered_cards = cards
    if project_filter:
        filtered_cards = [c for c in filtered_cards if c['project_id'] == int(project_filter)]
    if status_filter:
        filtered_cards = [c for c in filtered_cards if c['status'] == status_filter]
    if priority_filter:
        filtered_cards = [c for c in filtered_cards if c['priority'] == priority_filter]
    if assignee_filter:
        filtered_cards = [c for c in filtered_cards if c.get('assignee', '').lower() == assignee_filter.lower()]
    if search_query:
        filtered_cards = [c for c in filtered_cards if 
                         search_query in c['title'].lower() or 
                         search_query in c.get('description', '').lower()]
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues Export"
    
    # Headers
    headers = ['ID', 'Title', 'Description', 'Status', 'Priority', 'Assignee', 'Project', 'Due Date', 'Comments', 'Created Date']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0079BF', end_color='0079BF', fill_type='solid')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for card in filtered_cards:
        comment_count = get_comment_count(card['id'])
        project_name = projects.get(card['project_id'], 'Unknown')
        
        row_data = [
            f"#{card['id']}",
            card['title'],
            card.get('description', ''),
            card['status'].replace('_', ' ').title(),
            card['priority'],
            card.get('assignee', 'Unassigned'),
            project_name,
            card.get('due_date', ''),
            comment_count,
            card['created_at']
        ]
        ws.append(row_data)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with current date
    filename = f"issues_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    # Start email notification scheduler
    schedule_daily_notifications()
    app.run(debug=True)