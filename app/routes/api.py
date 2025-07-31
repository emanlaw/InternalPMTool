from flask import Blueprint, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
from flask_mail import Message
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from app.models.user import load_data, save_data
from app.utils.date_helpers import get_comment_count
from app.services.analytics_service import AnalyticsService

api_bp = Blueprint('api', __name__)

@api_bp.route('/test-email')
@login_required
def test_email():
    """Test email functionality"""
    try:
        data = load_data()
        user_data = next((u for u in data['users'] if u['id'] == current_user.id), None)
        
        if not user_data or not user_data.get('email'):
            return jsonify({'error': 'No email address set in profile'}), 400
        
        # Send test email
        msg = Message(
            subject="PM Tool - Test Email",
            recipients=[user_data['email']],
            html="""
            <h2>Test Email Successful! 🎉</h2>
            <p>Your email notifications are working correctly.</p>
            <p>You will receive daily notifications for overdue cards.</p>
            """
        )
        
        current_app.extensions['mail'].send(msg)
        return jsonify({'success': True, 'message': 'Test email sent successfully!'})
        
    except Exception as e:
        return jsonify({'error': f'Email test failed: {str(e)}'}), 500

@api_bp.route('/move_card', methods=['POST'])
def move_card():
    data = load_data()
    card_id = request.json['card_id']
    new_status = request.json['status']
    
    for card in data['cards']:
        if card['id'] == card_id:
            card['status'] = new_status
            break
    
    save_data(data)
    return jsonify({'success': True})

@api_bp.route('/add_card', methods=['POST'])
def add_card():
    try:
        from app.services.firebase_service import firebase_service
        from config.firebase_config import firebase_config
        
        if firebase_config.db is not None:
            # Use Firebase
            new_card = {
                'project_id': str(request.json['project_id']),
                'title': request.json['title'],
                'description': request.json.get('description', ''),
                'status': 'todo',
                'assignee': request.json.get('assignee', ''),
                'priority': request.json.get('priority', 'Medium'),
                'due_date': request.json.get('due_date', '')
            }
            created_card = firebase_service.create_card(new_card)
            return jsonify(created_card)
        else:
            return jsonify({'error': 'Firebase not configured. Please set up Firebase credentials.'}), 500
            
    except Exception as e:
        print(f"Firebase error: {e}")
        return jsonify({'error': f'Failed to create card: {str(e)}'}), 500

@api_bp.route('/add_project', methods=['POST'])
def add_project():
    try:
        from app.services.firebase_service import firebase_service
        from config.firebase_config import firebase_config
        
        if firebase_config.db is not None:
            # Use Firebase
            new_project = {
                'name': request.json['name'],
                'description': request.json.get('description', '')
            }
            created_project = firebase_service.create_project(new_project)
            return jsonify(created_project)
        else:
            return jsonify({'error': 'Firebase not configured. Please set up Firebase credentials.'}), 500
            
    except Exception as e:
        print(f"Firebase error: {e}")
        return jsonify({'error': f'Failed to create project: {str(e)}'}), 500

@api_bp.route('/update_card_status', methods=['POST'])
def update_card_status():
    try:
        from app.services.firebase_service import firebase_service
        from config.firebase_config import firebase_config
        
        card_id = request.json['card_id']
        new_status = request.json['status']
        
        if firebase_config.db is not None:
            # Get all cards to find the correct Firebase ID
            all_cards = firebase_service.get_all_cards()
            
            # Find the card by matching the ID (handle both integer and string formats)
            target_card = None
            for card in all_cards:
                # Check if card ID matches directly (for integer IDs)
                if card['id'] == card_id or str(card['id']) == str(card_id):
                    target_card = card
                    break
                # Also check CARD-XXX format
                elif isinstance(card['id'], str) and card['id'].startswith('CARD-'):
                    try:
                        card_number = int(card['id'].split('-')[1])
                        if card_number == card_id:
                            target_card = card
                            break
                    except (ValueError, IndexError):
                        continue
            
            if target_card:
                # Update the card status in Firebase
                success = firebase_service.update_card(str(target_card['id']), {'status': new_status})
                
                if success:
                    return jsonify({'success': True})
                else:
                    return jsonify({'error': 'Failed to update card status'}), 500
            else:
                return jsonify({'error': f'Card with ID {card_id} not found'}), 404
        else:
            return jsonify({'error': 'Firebase not configured'}), 500
            
    except Exception as e:
        print(f"Error updating card status: {e}")
        return jsonify({'error': f'Failed to update status: {str(e)}'}), 500

@api_bp.route('/card/<int:card_id>/comments')
@login_required
def get_comments(card_id):
    data = load_data()
    comments = [c for c in data.get('comments', []) if c['card_id'] == card_id]
    return jsonify(comments)

@api_bp.route('/card/<int:card_id>/comments', methods=['POST'])
@login_required
def add_comment(card_id):
    try:
        from app.services.firebase_service import firebase_service
        from config.firebase_config import firebase_config
        
        if firebase_config.db is not None:
            # Use Firebase
            new_comment = {
                'card_id': str(card_id),
                'author': current_user.username,
                'content': request.json['content']
            }
            created_comment = firebase_service.create_comment(new_comment)
            return jsonify(created_comment)
        else:
            return jsonify({'error': 'Firebase not configured. Please set up Firebase credentials.'}), 500
            
    except Exception as e:
        print(f"Firebase error: {e}")
        return jsonify({'error': f'Failed to create comment: {str(e)}'}), 500

@api_bp.route('/card/<int:card_id>')
@login_required
def get_card_details(card_id):
    data = load_data()
    card = next((c for c in data['cards'] if c['id'] == card_id), None)
    if not card:
        return jsonify({'error': 'Card not found'}), 404
    
    comments = [c for c in data.get('comments', []) if c['card_id'] == card_id]
    return jsonify({'card': card, 'comments': comments})

@api_bp.route('/search')
@login_required
def search_cards():
    data = load_data()
    cards = data['cards']
    
    # Get search parameters
    search_query = request.args.get('q', '').lower()
    status_filter = request.args.get('status', '')
    priority_filter = request.args.get('priority', '')
    assignee_filter = request.args.get('assignee', '')
    project_filter = request.args.get('project_id', '')
    
    # Filter cards
    filtered_cards = cards
    
    # Project filter
    if project_filter:
        filtered_cards = [c for c in filtered_cards if c['project_id'] == int(project_filter)]
    
    # Status filter
    if status_filter:
        filtered_cards = [c for c in filtered_cards if c['status'] == status_filter]
    
    # Priority filter
    if priority_filter:
        filtered_cards = [c for c in filtered_cards if c['priority'] == priority_filter]
    
    # Assignee filter
    if assignee_filter:
        filtered_cards = [c for c in filtered_cards if c.get('assignee', '').lower() == assignee_filter.lower()]
    
    # Search query (title and description)
    if search_query:
        filtered_cards = [c for c in filtered_cards if 
                         search_query in c['title'].lower() or 
                         search_query in c.get('description', '').lower()]
    
    return jsonify({
        'cards': filtered_cards,
        'total': len(filtered_cards)
    })

@api_bp.route('/export-excel')
@login_required
def export_excel():
    data = load_data()
    cards = data['cards']
    projects = {p['id']: p['name'] for p in data['projects']}
    
    # Apply same filters as search
    search_query = request.args.get('q', '').lower()
    status_filter = request.args.get('status', '')
    priority_filter = request.args.get('priority', '')
    assignee_filter = request.args.get('assignee', '')
    project_filter = request.args.get('project_id', '')
    
    # Filter cards (same logic as search)
    filtered_cards = cards
    if project_filter:
        filtered_cards = [c for c in filtered_cards if c['project_id'] == int(project_filter)]
    if status_filter:
        filtered_cards = [c for c in filtered_cards if c['status'] == status_filter]
    if priority_filter:
        filtered_cards = [c for c in filtered_cards if c['priority'] == priority_filter]
    if assignee_filter:
        filtered_cards = [c for c in filtered_cards if c.get('assignee', '').lower() == assignee_filter.lower()]
    if search_query:
        filtered_cards = [c for c in filtered_cards if 
                         search_query in c['title'].lower() or 
                         search_query in c.get('description', '').lower()]
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues Export"
    
    # Headers
    headers = ['ID', 'Title', 'Description', 'Status', 'Priority', 'Assignee', 'Project', 'Due Date', 'Comments', 'Created Date']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0079BF', end_color='0079BF', fill_type='solid')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for card in filtered_cards:
        comment_count = get_comment_count(card['id'])
        project_name = projects.get(card['project_id'], 'Unknown')
        
        row_data = [
            f"#{card['id']}",
            card['title'],
            card.get('description', ''),
            card['status'].replace('_', ' ').title(),
            card['priority'],
            card.get('assignee', 'Unassigned'),
            project_name,
            card.get('due_date', ''),
            comment_count,
            card['created_at']
        ]
        ws.append(row_data)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with current date
    filename = f"issues_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Analytics endpoints
@api_bp.route('/analytics/data')
@login_required
def get_analytics_data():
    """Get analytics data for dashboard"""
    analytics = AnalyticsService()
    
    return jsonify({
        'projects': analytics.get_project_health_data(),
        'resources': analytics.get_resource_utilization_data(),
        'performance': analytics.get_performance_metrics(),
        'predictive': analytics.get_predictive_analytics()
    })

@api_bp.route('/analytics/report/<report_type>')
@login_required
def generate_analytics_report(report_type):
    """Generate specific analytics report"""
    analytics = AnalyticsService()
    
    if report_type == 'project-health':
        report = analytics.generate_project_health_report()
    elif report_type == 'resource-utilization':
        report = analytics.generate_resource_utilization_report()
    elif report_type == 'performance-metrics':
        report = analytics.generate_performance_metrics_report()
    elif report_type == 'risk-assessment':
        report = analytics.generate_risk_assessment_report()
    else:
        return jsonify({'error': 'Invalid report type'}), 400
    
    return jsonify(report)

@api_bp.route('/debug/cards')
@login_required
def debug_cards():
    """Debug endpoint to check card ID mapping"""
    try:
        from app.services.firebase_service import firebase_service
        from config.firebase_config import firebase_config
        
        if firebase_config.db is not None:
            cards = firebase_service.get_all_cards()
            
            debug_info = []
            for card in cards:
                # Extract number from Firebase ID
                card_number = None
                if card['id'].startswith('CARD-'):
                    card_number = int(card['id'].split('-')[1])
                
                debug_info.append({
                    'firebase_id': card['id'],
                    'display_id': card_number,
                    'title': card['title'],
                    'status': card['status']
                })
            
            return jsonify({
                'success': True,
                'cards': debug_info
            })
        else:
            return jsonify({'error': 'Firebase not configured'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api_bp.route('/move_to_backlog', methods=['POST'])
@login_required
def move_to_backlog():
    """Move issue to backlog (todo status)"""
    try:
        from app.services.firebase_service import firebase_service
        from config.firebase_config import firebase_config
        
        card_id = request.json['card_id']
        
        if firebase_config.db is not None:
            # Get all cards to find the correct Firebase ID
            all_cards = firebase_service.get_all_cards()
            
            # Find the card by matching the ID
            target_card = None
            for card in all_cards:
                if card['id'] == card_id or str(card['id']) == str(card_id):
                    target_card = card
                    break
                elif isinstance(card['id'], str) and card['id'].startswith('CARD-'):
                    try:
                        card_number = int(card['id'].split('-')[1])
                        if card_number == card_id:
                            target_card = card
                            break
                    except (ValueError, IndexError):
                        continue
            
            if target_card:
                success = firebase_service.update_card(str(target_card['id']), {'status': 'todo'})
                
                if success:
                    return jsonify({'success': True})
                else:
                    return jsonify({'error': 'Failed to update card status'}), 500
            else:
                return jsonify({'error': f'Card with ID {card_id} not found'}), 404
        else:
            return jsonify({'error': 'Firebase not configured'}), 500
            
    except Exception as e:
        print(f"Error moving to backlog: {e}")
        return jsonify({'error': f'Failed to move to backlog: {str(e)}'}), 500

@api_bp.route('/send_to_assignee', methods=['POST'])
@login_required
def send_to_assignee():
    """Send email notification to assignee"""
    try:
        from app.services.firebase_service import firebase_service
        from config.firebase_config import firebase_config
        
        card_id = request.json['card_id']
        assignee = request.json['assignee']
        
        if not assignee or assignee == 'Unassigned':
            return jsonify({'error': 'No assignee specified'}), 400
        
        # Get card details
        if firebase_config.db is not None:
            all_cards = firebase_service.get_all_cards()
            target_card = None
            
            for card in all_cards:
                if card['id'] == card_id or str(card['id']) == str(card_id):
                    target_card = card
                    break
                elif isinstance(card['id'], str) and card['id'].startswith('CARD-'):
                    try:
                        card_number = int(card['id'].split('-')[1])
                        if card_number == card_id:
                            target_card = card
                            break
                    except (ValueError, IndexError):
                        continue
            
            if not target_card:
                return jsonify({'error': 'Card not found'}), 404
            
            # Send email notification
            try:
                msg = Message(
                    subject=f"PM Tool - Issue Assigned: {target_card['title']}",
                    recipients=[f"{assignee}@company.com"],  # Assuming email format
                    html=f"""
                    <h2>Issue Assigned to You</h2>
                    <p><strong>Title:</strong> {target_card['title']}</p>
                    <p><strong>Description:</strong> {target_card.get('description', 'No description')}</p>
                    <p><strong>Priority:</strong> {target_card['priority']}</p>
                    <p><strong>Status:</strong> {target_card['status'].replace('_', ' ').title()}</p>
                    <p><strong>Due Date:</strong> {target_card.get('due_date', 'Not set')}</p>
                    <br>
                    <p>Please check the PM Tool for more details.</p>
                    """
                )
                
                current_app.extensions['mail'].send(msg)
                return jsonify({'success': True, 'message': f'Email sent to {assignee}'})
                
            except Exception as e:
                print(f"Email error: {e}")
                return jsonify({'error': f'Failed to send email: {str(e)}'}), 500
        else:
            return jsonify({'error': 'Firebase not configured'}), 500
            
    except Exception as e:
        print(f"Error sending to assignee: {e}")
        return jsonify({'error': f'Failed to send to assignee: {str(e)}'}), 500

@api_bp.route('/delete_card', methods=['POST'])
@login_required
def delete_card():
    """Delete a card permanently"""
    try:
        from app.services.firebase_service import firebase_service
        from config.firebase_config import firebase_config
        
        card_id = request.json['card_id']
        
        if firebase_config.db is not None:
            # Get all cards to find the correct Firebase ID
            all_cards = firebase_service.get_all_cards()
            
            target_card = None
            for card in all_cards:
                if card['id'] == card_id or str(card['id']) == str(card_id):
                    target_card = card
                    break
                elif isinstance(card['id'], str) and card['id'].startswith('CARD-'):
                    try:
                        card_number = int(card['id'].split('-')[1])
                        if card_number == card_id:
                            target_card = card
                            break
                    except (ValueError, IndexError):
                        continue
            
            if target_card:
                # Delete the card from Firebase
                success = firebase_service.delete_card(str(target_card['id']))
                
                if success:
                    return jsonify({'success': True})
                else:
                    return jsonify({'error': 'Failed to delete card from database'}), 500
            else:
                return jsonify({'error': f'Card with ID {card_id} not found'}), 404
        else:
            return jsonify({'error': 'Firebase not configured'}), 500
            
    except Exception as e:
        print(f"Error deleting card: {e}")
        return jsonify({'error': f'Failed to delete card: {str(e)}'}), 500